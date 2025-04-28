import os
import decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import date

def get_reading_level(data):
    score = 0
    for i in range(len(data)):
        score = score + data[i][0]

    return round((decimal.Decimal(score) / decimal.Decimal(20)), 1) 

def create_excel(student):

    if not os.path.isdir("ReadingAssessmentResults"):

        os.makedirs("ReadingAssessmentResults")

    data = student.get_result()
    reading_level = get_reading_level(data)
    name = student.get_firstname() + ", " + student.get_lastname()
    title = f"{reading_level}_" + student.get_lastname().upper() + student.get_firstname().replace(" ","")
    headings = ["ListP", "List1", "List2", "List3", "List4", "List5", "List6", "List7", "List8", "ListHS"]

    wb = Workbook()
    ws = wb.active
    ws.title = title

    ws.merge_cells('A2:E2')
    ws['A2'] = "Name: " + name
    ws.merge_cells('G2:J2')
    ws['G2'] = "Reading Level: " + str(reading_level)
    ws.merge_cells('A3:E3')
    ws['A3'] = "School: " + student.get_school()
    ws.merge_cells('A4:E4')
    ws['A4'] = "Instructor: " + student.get_instructor()
    ws.merge_cells('G4:J4')
    ws['G4'] = "Instructor ID: " + student.get_instructorID()
    col = 1
    col2 = 1
    for i in range(len(headings)):
        row = 6
        if i < 5:        
            ws.merge_cells(get_column_letter(col)+str(row)+":"+get_column_letter(col+1)+str(row))
            ws.merge_cells(get_column_letter(col)+str(row+1)+":"+get_column_letter(col+1)+str(row+1))
            ws[get_column_letter(col)+str(row)] = headings[i]
            ws[get_column_letter(col)+str(row+1)] = data[i][0]
            for x in range(1, 20):
                ws[get_column_letter(col)+str(row + x + 1)] = data[i][1][x-1]
                ws[get_column_letter(col + 1)+str(row + x + 1)] = data[i][2][x-1]
            col = col + 2
        else:
            row = row + 22
            ws.merge_cells(get_column_letter(col2)+str(row)+":"+get_column_letter(col2+1)+str(row))
            ws.merge_cells(get_column_letter(col2)+str(row+1)+":"+get_column_letter(col2+1)+str(row+1))
            ws[get_column_letter(col2)+str(row)] = headings[i]
            ws[get_column_letter(col2)+str(row+1)] = data[i][0]
            for x in range(1, 20):
                ws[get_column_letter(col2)+str(row + x + 1)] = data[i][1][x-1]
                ws[get_column_letter(col2 + 1)+str(row + x + 1)] = data[i][2][x-1]
            col2 = col2 + 2


    wb.save("ReadingAssessmentResults/" + title + str(date.today()) + ".xlsx")